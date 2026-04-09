import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

PATH = "/home/zo/University/FA583/Admin/RELX_Financial_Analysis_FA583_FINAL.xlsx"
wb = openpyxl.load_workbook(PATH)

GREY_HDR = PatternFill("solid", fgColor="F2F2F2")
NO_FILL  = PatternFill(fill_type=None)

def thin(color="BFBFBF"):
    return Side(style="thin", color=color)

BORDER_ALL = Border(left=thin(), right=thin(), top=thin(), bottom=thin())
BORDER_LRB = Border(left=thin(), right=thin(), bottom=thin())
BORDER_TOP = Border(top=thin())
BORDER_TB  = Border(top=thin(), bottom=thin())
BORDER_BOT_MED = Border(bottom=Side(style="medium", color="2F2F2F"))
BORDER_NONE = Border()

FONT_TITLE  = Font(name="Calibri", bold=True, size=12, color="000000")
FONT_HDR    = Font(name="Calibri", bold=True, size=10, color="000000")
FONT_SOURCE = Font(name="Calibri", italic=True, size=9, color="595959")
FONT_BODY   = Font(name="Calibri", size=10, color="000000")
FONT_BOLD   = Font(name="Calibri", bold=True, size=10, color="000000")
FONT_ITALIC = Font(name="Calibri", italic=True, size=10, color="000000")

ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Track merged cell ranges to avoid double-styling
def get_merged_top_lefts(ws):
    top_lefts = set()
    for merged in ws.merged_cells.ranges:
        top_lefts.add((merged.min_row, merged.min_col))
    return top_lefts

def is_in_merged_not_top_left(ws, row, col):
    for merged in ws.merged_cells.ranges:
        if (merged.min_row <= row <= merged.max_row and
            merged.min_col <= col <= merged.max_col):
            if not (row == merged.min_row and col == merged.min_col):
                return True
    return False

def style_ws(ws):
    max_col = ws.max_column
    for row_idx in range(1, ws.max_row + 1):
        row_cells = list(ws.iter_rows(min_row=row_idx, max_row=row_idx,
                                       min_col=1, max_col=max_col))[0]

        col_a = row_cells[0].value
        col_b = row_cells[1].value if max_col > 1 else None
        col_c = row_cells[2].value if max_col > 2 else None
        col_d = row_cells[3].value if max_col > 3 else None
        col_a_str = str(col_a).strip() if col_a is not None else ""

        is_blank      = all(c.value is None for c in row_cells)
        is_title      = (row_idx == 1)
        is_source     = (row_idx == 2)
        is_col_header = (row_idx == 4)
        is_section    = (row_idx >= 5 and col_a is not None and
                         col_b is None and col_c is None and col_d is None and
                         col_a_str != "")
        is_memo       = col_a_str.upper().startswith("MEMO")
        is_total      = any(col_a_str.lower().startswith(x) for x in [
                            "total", "net cash", "gross profit",
                            "profit before tax", "net profit (total)"])

        for cell in row_cells:
            # Skip non-top-left merged cells (don't re-style them)
            if is_in_merged_not_top_left(ws, cell.row, cell.column):
                continue

            if is_blank:
                cell.fill   = NO_FILL
                cell.border = BORDER_NONE
                cell.font   = FONT_BODY
                continue

            if is_title:
                cell.fill      = NO_FILL
                cell.font      = FONT_TITLE
                cell.border    = BORDER_BOT_MED
                cell.alignment = ALIGN_LEFT
                continue

            if is_source:
                cell.fill      = NO_FILL
                cell.font      = FONT_SOURCE
                cell.border    = BORDER_NONE
                cell.alignment = ALIGN_LEFT
                continue

            if is_col_header:
                cell.fill      = GREY_HDR
                cell.font      = FONT_HDR
                cell.border    = BORDER_ALL
                cell.alignment = ALIGN_LEFT if cell.column == 1 else ALIGN_CENTER
                continue

            if is_section:
                cell.fill      = NO_FILL
                cell.font      = FONT_HDR
                cell.border    = BORDER_TOP
                cell.alignment = ALIGN_LEFT
                continue

            if is_memo:
                cell.fill      = NO_FILL
                cell.font      = FONT_ITALIC
                cell.border    = BORDER_TB
                cell.alignment = ALIGN_LEFT if cell.column == 1 else ALIGN_CENTER
                continue

            if is_total:
                cell.fill      = NO_FILL
                cell.font      = FONT_BOLD
                cell.border    = BORDER_ALL
                cell.alignment = ALIGN_LEFT if cell.column == 1 else ALIGN_CENTER
                continue

            # Regular data row
            cell.fill      = NO_FILL
            cell.font      = FONT_BODY
            cell.border    = BORDER_LRB
            cell.alignment = ALIGN_LEFT if cell.column == 1 else ALIGN_CENTER

for sheet_name in wb.sheetnames:
    style_ws(wb[sheet_name])
    print(f"Styled: {sheet_name}")

wb.save(PATH)
print("Saved:", PATH)
