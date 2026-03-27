import openpyxl

PATH = "coursework/RELX_Financial_Analysis_FA583_FINAL.xlsx"
wb = openpyxl.load_workbook(PATH)

# --- Fix 1: Delete Notes column (col 6) from Ratios sheet ---
ws_ratios = wb["Ratios"]
ws_ratios.delete_cols(6)  # col F = Notes

# --- Fix 2: Delete MEMO row from Balance Sheet ---
ws_bs = wb["Balance Sheet"]
for row in list(ws_bs.iter_rows()):
    for cell in row:
        if cell.value and str(cell.value).startswith("MEMO: Net debt"):
            ws_bs.delete_rows(cell.row)
            break

# --- Fix 3: Delete MEMO row from Cash Flow ---
ws_cf = wb["Cash Flow"]
for row in list(ws_cf.iter_rows()):
    for cell in row:
        if cell.value and str(cell.value).startswith("MEMO: Free cash flow"):
            ws_cf.delete_rows(cell.row)
            break

wb.save(PATH)
print("Done. Verify the Ratios sheet now has 5 columns, and no MEMO rows exist.")
