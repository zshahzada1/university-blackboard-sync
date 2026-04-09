#!/usr/bin/env python3
"""
Rebuild the Ratios sheet in RELX_Financial_Analysis_FA583_FINAL.xlsx
with 8 ratios from Melville's textbook Ch.22.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# File path
FILE_PATH = '/home/zo/University/FA583/Admin/RELX_Financial_Analysis_FA583_FINAL.xlsx'

# Load workbook
print("Loading workbook...")
wb = openpyxl.load_workbook(FILE_PATH)
ws_income = wb['Income Statement']
ws_balance = wb['Balance Sheet']

# Extract source values from Income Statement (rows are 1-indexed)
revenue = [ws_income[f'{col}5'].value for col in ['B', 'C', 'D']]
cost_of_sales = [abs(ws_income[f'{col}6'].value) for col in ['B', 'C', 'D']]
gross_profit = [ws_income[f'{col}7'].value for col in ['B', 'C', 'D']]
ebit = [ws_income[f'{col}11'].value for col in ['B', 'C', 'D']]
net_profit = [ws_income[f'{col}19'].value for col in ['B', 'C', 'D']]

# Extract source values from Balance Sheet
inventories = [ws_balance[f'{col}12'].value for col in ['B', 'C', 'D']]
trade_receivables = [ws_balance[f'{col}13'].value for col in ['B', 'C', 'D']]
current_assets = [ws_balance[f'{col}16'].value for col in ['B', 'C', 'D']]
total_assets = [ws_balance[f'{col}18'].value for col in ['B', 'C', 'D']]
current_liabilities = [ws_balance[f'{col}24'].value for col in ['B', 'C', 'D']]
long_term_borrowings = [ws_balance[f'{col}27'].value for col in ['B', 'C', 'D']]
equity = [ws_balance[f'{col}34'].value for col in ['B', 'C', 'D']]

# Calculate ratios for each year
print("Calculating ratios...")
ratios_data = []

# Validate extracted values before calculating
required = {
    "ebit": ebit, "revenue": revenue, "cost_of_sales": cost_of_sales,
    "gross_profit": gross_profit, "net_profit": net_profit,
    "inventories": inventories, "trade_receivables": trade_receivables,
    "current_assets": current_assets, "total_assets": total_assets,
    "current_liabilities": current_liabilities,
    "long_term_borrowings": long_term_borrowings, "equity": equity,
}
for year_idx, year in enumerate([2022, 2023, 2024]):
    for name, values in required.items():
        if values[year_idx] is None:
            raise ValueError(f"{year}: {name} is None — check source sheet row/column references")
        if not isinstance(values[year_idx], (int, float)):
            raise ValueError(f"{year}: {name} = {values[year_idx]!r} is not numeric")

for i in range(3):
    # 1. ROCE = EBIT / (Total assets − Current liabilities)
    roce = ebit[i] / (total_assets[i] - current_liabilities[i])

    # 2. Gross profit margin
    gp_margin = gross_profit[i] / revenue[i]

    # 3. Net profit margin
    np_margin = net_profit[i] / revenue[i]

    # 4. Current ratio
    current_ratio = current_assets[i] / current_liabilities[i]

    # 5. Quick assets ratio
    quick = (current_assets[i] - inventories[i]) / current_liabilities[i]

    # 6. Inventory holding period (in days)
    inv_period = (inventories[i] / cost_of_sales[i]) * 365

    # 7. Trade receivables collection period (in days)
    tr_period = (trade_receivables[i] / revenue[i]) * 365

    # 8. Capital gearing ratio
    gearing = long_term_borrowings[i] / (long_term_borrowings[i] + equity[i])

    ratios_data.append({
        'roce': roce,
        'gp_margin': gp_margin,
        'np_margin': np_margin,
        'current_ratio': current_ratio,
        'quick_ratio': quick,
        'inv_period': inv_period,
        'tr_period': tr_period,
        'gearing': gearing
    })

# Prepare Ratios sheet
print("Rebuilding Ratios sheet...")
ws_ratios = wb['Ratios']

# Clear the sheet (delete all rows except the first one, then delete and recreate)
ws_ratios.delete_rows(1, ws_ratios.max_row)

# Define colors and styles
navy_fill = PatternFill(start_color='1F3864', end_color='1F3864', fill_type='solid')
mid_blue_fill = PatternFill(start_color='2E5496', end_color='2E5496', fill_type='solid')
light_grey_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

white_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
white_bold_11 = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
grey_font_9 = Font(name='Calibri', size=9, italic=True, color='808080')
black_font_10 = Font(name='Calibri', size=10)

thin_border = Border(
    left=Side(style='thin', color='BFBFBF'),
    right=Side(style='thin', color='BFBFBF'),
    top=Side(style='thin', color='BFBFBF'),
    bottom=Side(style='thin', color='BFBFBF')
)

# Row 1: Title
ws_ratios.merge_cells('A1:F1')
cell_a1 = ws_ratios['A1']
cell_a1.value = 'RELX PLC — Financial Ratio Analysis (£m)'
cell_a1.font = white_font
cell_a1.fill = navy_fill
cell_a1.alignment = Alignment(horizontal='center', vertical='center')
for col in ['A', 'B', 'C', 'D', 'E', 'F']:
    ws_ratios[f'{col}1'].border = thin_border
    ws_ratios[f'{col}1'].fill = navy_fill
ws_ratios.row_dimensions[1].height = 22

# Row 2: Source note
ws_ratios['A2'].value = 'Source: Melville, A. (2019) International Financial Reporting, 7th ed., Ch.22 | Data: RELX Annual Reports 2022–2024'
ws_ratios['A2'].font = grey_font_9
ws_ratios['A2'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
for col in ['A', 'B', 'C', 'D', 'E', 'F']:
    ws_ratios[f'{col}2'].border = thin_border
ws_ratios.row_dimensions[2].height = 14

# Row 3: Column headers
headers = ['Ratio', '2022', '2023', '2024', 'Formula', 'Notes']
for col_idx, header in enumerate(headers, start=1):
    col_letter = get_column_letter(col_idx)
    cell = ws_ratios[f'{col_letter}3']
    cell.value = header
    cell.font = white_bold_11
    cell.fill = navy_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border
ws_ratios.row_dimensions[3].height = 18

# Row 4: Section header - PROFITABILITY
ws_ratios.merge_cells('A4:F4')
cell_a4 = ws_ratios['A4']
cell_a4.value = 'PROFITABILITY'
cell_a4.font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
cell_a4.fill = mid_blue_fill
cell_a4.alignment = Alignment(horizontal='left', vertical='center')
for col in ['A', 'B', 'C', 'D', 'E', 'F']:
    ws_ratios[f'{col}4'].border = thin_border
    ws_ratios[f'{col}4'].fill = mid_blue_fill
ws_ratios.row_dimensions[4].height = 16

# Ratio data rows (starting from row 5)
ratio_rows = [
    {
        'name': 'ROCE',
        'formula': 'EBIT / (Total assets − Current liabilities)',
        'note': 'Capital employed = total assets less current liabilities',
        'values': [ratios_data[i]['roce'] for i in range(3)],
        'format': '0.0%'
    },
    {
        'name': 'Gross profit margin',
        'formula': 'Gross profit / Revenue',
        'note': 'Stable ~65% — subscription cost base is predictable',
        'values': [ratios_data[i]['gp_margin'] for i in range(3)],
        'format': '0.0%'
    },
    {
        'name': 'Net profit margin',
        'formula': 'Net profit (shareholders) / Revenue',
        'note': 'Improved despite UK tax rise to 25% (Apr 2023)',
        'values': [ratios_data[i]['np_margin'] for i in range(3)],
        'format': '0.0%'
    },
]

row_num = 5
for idx, ratio in enumerate(ratio_rows):
    # Determine alternating fill
    fill_color = light_grey_fill if idx % 2 == 0 else white_fill

    # Column A: Ratio name
    ws_ratios[f'A{row_num}'].value = ratio['name']
    ws_ratios[f'A{row_num}'].font = black_font_10
    ws_ratios[f'A{row_num}'].fill = fill_color
    ws_ratios[f'A{row_num}'].alignment = Alignment(horizontal='left', vertical='center')
    ws_ratios[f'A{row_num}'].border = thin_border

    # Columns B, C, D: Values
    for col_idx, col_letter in enumerate(['B', 'C', 'D']):
        cell = ws_ratios[f'{col_letter}{row_num}']
        cell.value = ratio['values'][col_idx]
        cell.number_format = ratio['format']
        cell.fill = fill_color
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # Column E: Formula
    ws_ratios[f'E{row_num}'].value = ratio['formula']
    ws_ratios[f'E{row_num}'].font = black_font_10
    ws_ratios[f'E{row_num}'].fill = fill_color
    ws_ratios[f'E{row_num}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws_ratios[f'E{row_num}'].border = thin_border

    # Column F: Notes
    ws_ratios[f'F{row_num}'].value = ratio['note']
    ws_ratios[f'F{row_num}'].font = black_font_10
    ws_ratios[f'F{row_num}'].fill = fill_color
    ws_ratios[f'F{row_num}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws_ratios[f'F{row_num}'].border = thin_border

    ws_ratios.row_dimensions[row_num].height = 16
    row_num += 1

# Row 8: Section header - LIQUIDITY SHORT-TERM
ws_ratios.merge_cells('A8:F8')
cell_a8 = ws_ratios['A8']
cell_a8.value = 'LIQUIDITY — SHORT-TERM'
cell_a8.font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
cell_a8.fill = mid_blue_fill
cell_a8.alignment = Alignment(horizontal='left', vertical='center')
for col in ['A', 'B', 'C', 'D', 'E', 'F']:
    ws_ratios[f'{col}8'].border = thin_border
    ws_ratios[f'{col}8'].fill = mid_blue_fill
ws_ratios.row_dimensions[8].height = 16

# Short-term liquidity ratios
liquidity_short_rows = [
    {
        'name': 'Current ratio',
        'formula': 'Current assets / Current liabilities',
        'note': 'Below 1.0x — structural: ~£4bn deferred subscription income in CL',
        'values': [ratios_data[i]['current_ratio'] for i in range(3)],
        'format': '0.00"x"'
    },
    {
        'name': 'Quick assets ratio (acid test)',
        'formula': '(Current assets − Inventories) / Current liabilities',
        'note': 'Closely tracks current ratio — minimal inventory effect',
        'values': [ratios_data[i]['quick_ratio'] for i in range(3)],
        'format': '0.00"x"'
    },
    {
        'name': 'Inventory holding period',
        'formula': '(Inventories / Cost of sales) × 365',
        'note': 'Stable ~37 days — mainly physical exhibition/print inventory',
        'values': [ratios_data[i]['inv_period'] for i in range(3)],
        'format': '0.0" days"'
    },
    {
        'name': 'Trade receivables collection period',
        'formula': '(Trade receivables / Revenue) × 365',
        'note': 'Long but improving — B2B enterprise clients on annual contracts',
        'values': [ratios_data[i]['tr_period'] for i in range(3)],
        'format': '0.0" days"'
    },
]

row_num = 9
for idx, ratio in enumerate(liquidity_short_rows):
    # Determine alternating fill
    fill_color = light_grey_fill if idx % 2 == 0 else white_fill

    # Column A: Ratio name
    ws_ratios[f'A{row_num}'].value = ratio['name']
    ws_ratios[f'A{row_num}'].font = black_font_10
    ws_ratios[f'A{row_num}'].fill = fill_color
    ws_ratios[f'A{row_num}'].alignment = Alignment(horizontal='left', vertical='center')
    ws_ratios[f'A{row_num}'].border = thin_border

    # Columns B, C, D: Values
    for col_idx, col_letter in enumerate(['B', 'C', 'D']):
        cell = ws_ratios[f'{col_letter}{row_num}']
        cell.value = ratio['values'][col_idx]
        cell.number_format = ratio['format']
        cell.fill = fill_color
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # Column E: Formula
    ws_ratios[f'E{row_num}'].value = ratio['formula']
    ws_ratios[f'E{row_num}'].font = black_font_10
    ws_ratios[f'E{row_num}'].fill = fill_color
    ws_ratios[f'E{row_num}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws_ratios[f'E{row_num}'].border = thin_border

    # Column F: Notes
    ws_ratios[f'F{row_num}'].value = ratio['note']
    ws_ratios[f'F{row_num}'].font = black_font_10
    ws_ratios[f'F{row_num}'].fill = fill_color
    ws_ratios[f'F{row_num}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws_ratios[f'F{row_num}'].border = thin_border

    ws_ratios.row_dimensions[row_num].height = 16
    row_num += 1

# Row 13: Section header - LIQUIDITY LONG-TERM / SOLVENCY
ws_ratios.merge_cells('A13:F13')
cell_a13 = ws_ratios['A13']
cell_a13.value = 'LIQUIDITY — LONG-TERM / SOLVENCY'
cell_a13.font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
cell_a13.fill = mid_blue_fill
cell_a13.alignment = Alignment(horizontal='left', vertical='center')
for col in ['A', 'B', 'C', 'D', 'E', 'F']:
    ws_ratios[f'{col}13'].border = thin_border
    ws_ratios[f'{col}13'].fill = mid_blue_fill
ws_ratios.row_dimensions[13].height = 16

# Long-term solvency ratio
liquidity_long_rows = [
    {
        'name': 'Capital gearing ratio',
        'formula': 'Long-term borrowings / (Long-term borrowings + Equity)',
        'note': 'Elevated but slowly declining; equity base partly intangible',
        'values': [ratios_data[i]['gearing'] for i in range(3)],
        'format': '0.0%'
    },
]

row_num = 14
for idx, ratio in enumerate(liquidity_long_rows):
    # Determine alternating fill
    fill_color = light_grey_fill if idx % 2 == 0 else white_fill

    # Column A: Ratio name
    ws_ratios[f'A{row_num}'].value = ratio['name']
    ws_ratios[f'A{row_num}'].font = black_font_10
    ws_ratios[f'A{row_num}'].fill = fill_color
    ws_ratios[f'A{row_num}'].alignment = Alignment(horizontal='left', vertical='center')
    ws_ratios[f'A{row_num}'].border = thin_border

    # Columns B, C, D: Values
    for col_idx, col_letter in enumerate(['B', 'C', 'D']):
        cell = ws_ratios[f'{col_letter}{row_num}']
        cell.value = ratio['values'][col_idx]
        cell.number_format = ratio['format']
        cell.fill = fill_color
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # Column E: Formula
    ws_ratios[f'E{row_num}'].value = ratio['formula']
    ws_ratios[f'E{row_num}'].font = black_font_10
    ws_ratios[f'E{row_num}'].fill = fill_color
    ws_ratios[f'E{row_num}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws_ratios[f'E{row_num}'].border = thin_border

    # Column F: Notes
    ws_ratios[f'F{row_num}'].value = ratio['note']
    ws_ratios[f'F{row_num}'].font = black_font_10
    ws_ratios[f'F{row_num}'].fill = fill_color
    ws_ratios[f'F{row_num}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws_ratios[f'F{row_num}'].border = thin_border

    ws_ratios.row_dimensions[row_num].height = 16
    row_num += 1

# Set column widths
ws_ratios.column_dimensions['A'].width = 36
ws_ratios.column_dimensions['B'].width = 10
ws_ratios.column_dimensions['C'].width = 10
ws_ratios.column_dimensions['D'].width = 10
ws_ratios.column_dimensions['E'].width = 44
ws_ratios.column_dimensions['F'].width = 50

# Freeze panes at A4
ws_ratios.freeze_panes = 'A4'

# Save the workbook
print("Saving workbook...")
wb.save("/home/zo/University/FA583/Admin/RELX_Financial_Analysis_FA583_FINAL.xlsx")
print(f"Successfully updated {FILE_PATH}")
