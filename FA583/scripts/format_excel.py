#!/usr/bin/env python3
"""
Format Income Statement, Balance Sheet, and Cash Flow sheets in Excel workbook.
Applies consistent styling per FA583 Task 2 requirements.
Preserves existing Ratios sheet styling.
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# File path
WORKBOOK_PATH = "/home/zo/University/FA583/Admin/RELX_Financial_Analysis_FA583_FINAL.xlsx"

# Color definitions (hex without #)
COLOR_NAVY = "1F3864"
COLOR_MID_BLUE = "2E5496"
COLOR_LIGHT_BLUE_GREY = "D9E1F2"
COLOR_GREY = "595959"
COLOR_BORDER = "BFBFBF"

# Sheets to format (skip Ratios)
SHEETS_TO_FORMAT = ["Income Statement", "Balance Sheet", "Cash Flow"]

# Column widths by sheet
COLUMN_WIDTHS = {
    "Income Statement": {"A": 36, "B": 12, "C": 12, "D": 12, "E": 30},
    "Balance Sheet": {"A": 36, "B": 12, "C": 12, "D": 12, "E": 30},
    "Cash Flow": {"A": 40, "B": 12, "C": 12, "D": 12},
}


def load_workbook():
    """Load the workbook."""
    return openpyxl.load_workbook(WORKBOOK_PATH)


def is_section_header(ws, row_num):
    """
    Check if a row is a section header.
    Section headers are rows where column B value is None/empty
    and column A has text (these are subtotals/section labels).
    """
    cell_b = ws.cell(row_num, 2)  # Column B
    cell_a = ws.cell(row_num, 1)  # Column A

    # Check if B is empty and A has text
    b_value = cell_b.value
    a_value = cell_a.value

    return (b_value is None or b_value == "") and a_value is not None


def format_title_row(ws, row_num):
    """Format row 1 (title row)."""
    fill = PatternFill(start_color=COLOR_NAVY, end_color=COLOR_NAVY, fill_type="solid")
    font = Font(bold=True, color="FFFFFF", size=12)

    # Apply to all used columns
    for col_num in range(1, ws.max_column + 1):
        cell = ws.cell(row_num, col_num)
        cell.fill = fill
        cell.font = font

    ws.row_dimensions[row_num].height = 22


def format_source_row(ws, row_num):
    """Format row 2 (source row)."""
    font = Font(italic=True, color=COLOR_GREY, size=9)

    # Apply to column A only (where text is)
    cell = ws.cell(row_num, 1)
    cell.font = font

    ws.row_dimensions[row_num].height = 14


def format_header_row(ws, row_num):
    """Format row 4 (column header row)."""
    fill = PatternFill(start_color=COLOR_NAVY, end_color=COLOR_NAVY, fill_type="solid")
    font = Font(bold=True, color="FFFFFF", size=11)
    alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style="thin", color=COLOR_BORDER),
        right=Side(style="thin", color=COLOR_BORDER),
        top=Side(style="thin", color=COLOR_BORDER),
        bottom=Side(style="thin", color=COLOR_BORDER),
    )

    for col_num in range(1, ws.max_column + 1):
        cell = ws.cell(row_num, col_num)
        cell.fill = fill
        cell.font = font
        cell.alignment = alignment
        cell.border = border

    ws.row_dimensions[row_num].height = 18


def format_data_rows(ws):
    """Format data rows (row 5 onwards) with alternating colors."""
    border = Border(
        left=Side(style="thin", color=COLOR_BORDER),
        right=Side(style="thin", color=COLOR_BORDER),
        top=Side(style="thin", color=COLOR_BORDER),
        bottom=Side(style="thin", color=COLOR_BORDER),
    )

    data_row_index = 0  # 0-indexed from row 5

    for row_num in range(5, ws.max_row + 1):
        # Determine if this is a section header
        if is_section_header(ws, row_num):
            # Section header styling
            fill = PatternFill(start_color=COLOR_MID_BLUE, end_color=COLOR_MID_BLUE, fill_type="solid")
            font = Font(bold=True, color="FFFFFF", size=10)

            for col_num in range(1, ws.max_column + 1):
                cell = ws.cell(row_num, col_num)
                cell.fill = fill
                cell.font = font
                cell.border = border

            ws.row_dimensions[row_num].height = 15
            data_row_index += 1  # Keep index incremented
        else:
            # Regular data row with alternating colors
            is_even_row = (data_row_index % 2) == 0

            if is_even_row:
                fill = PatternFill(start_color=COLOR_LIGHT_BLUE_GREY, end_color=COLOR_LIGHT_BLUE_GREY, fill_type="solid")
            else:
                fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

            font = Font(name="Calibri", size=10)

            for col_num in range(1, ws.max_column + 1):
                cell = ws.cell(row_num, col_num)
                cell.fill = fill
                cell.font = font
                cell.border = border

                # Column A: left-aligned; others: center-aligned
                if col_num == 1:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                    # Notes column (E): keep left-aligned, no number format
                elif col_num == 5:  # Column E (Notes)
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    # Apply number format to value columns (B, C, D)
                    if col_num in [2, 3, 4]:
                        cell.number_format = "#,##0"

            ws.row_dimensions[row_num].height = 15
            data_row_index += 1


def set_column_widths(ws, sheet_name):
    """Set column widths for the sheet."""
    if sheet_name not in COLUMN_WIDTHS:
        return

    widths = COLUMN_WIDTHS[sheet_name]
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


def freeze_panes(ws):
    """Freeze panes at B5."""
    ws.freeze_panes = "B5"


def format_sheet(wb, sheet_name):
    """Format a single sheet."""
    ws = wb[sheet_name]

    # Format rows 1-4
    format_title_row(ws, 1)
    format_source_row(ws, 2)
    # Row 3 is blank, skip it
    format_header_row(ws, 4)

    # Format data rows
    format_data_rows(ws)

    # Set column widths
    set_column_widths(ws, sheet_name)

    # Freeze panes
    freeze_panes(ws)


def main():
    """Main function."""
    try:
        # Load workbook
        wb = load_workbook()

        # Format each sheet
        for sheet_name in SHEETS_TO_FORMAT:
            format_sheet(wb, sheet_name)

        # Save workbook
        wb.save(WORKBOOK_PATH)
        print("All sheets formatted OK")

    except Exception as e:
        print(f"Error: {e}")
        raise


if __name__ == "__main__":
    main()
